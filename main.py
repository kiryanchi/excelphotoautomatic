import sys
import openpyxl
from PyQt5.QtWidgets import QFileDialog, QApplication, QWidget, QLabel, QLayout, QHeaderView, QTableWidget, QVBoxLayout, \
    QHBoxLayout, QAbstractItemView, QMessageBox, QShortcut
from PyQt5.QtGui import QPixmap
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5.QtCore import QFile, QIODevice, QDataStream, QVariant
from PyQt5.QtGui import QKeySequence
import threading
from image.insert import insertinexcel
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SAVE_DIR = BASE_DIR + '\\Save'
form_class = uic.loadUiType(BASE_DIR + '\\UI\\' + "my_form.ui")[0]
FILE_NAME = 'default'
NOT_SAVE = False


def setLabelText(label, text):
    label.setText(text)
    label.adjustSize()


class TableWidgetPixmap(QPixmap):
    def __init__(self, imgpath):
        super().__init__()
        self.load(imgpath)


class TableWidget(QWidget):
    def __init__(self, imgpath, pixmap=None):
        super().__init__()
        tab_idx = myWindow.sheetlist.currentIndex()
        self.imgpath = imgpath
        self.lbl = QLabel()
        self.img = TableWidgetPixmap(imgpath) if pixmap is None else pixmap
        self.img = self.img.scaled(myWindow.sheetlist.widget(tab_idx).table.width() // 3 - 30, 190 + 10)
        self.lbl.setPixmap(self.img)
        self.widget_layout = QHBoxLayout()
        self.widget_layout.addWidget(self.lbl)
        self.widget_layout.setSizeConstraint(QLayout.SetFixedSize)
        self.setLayout(self.widget_layout)


class MyTable(QTableWidget):
    def __init__(self):
        super().__init__(0, 3)
        self.setAcceptDrops(True)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        # self.setDragDropMode(QAbstractItemView.InternalMove)
        # self.setDefaultDropAction(QtCore.Qt.CopyAction)
        # 테이블위젯의 헤더 크기를 조정
        self.setAlternatingRowColors(True)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setMinimumSize(QtCore.QSize(990, 630))
        self.verticalHeader().setDefaultSectionSize(220)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.setHorizontalHeaderLabels(['작업 전', '작업 후', '전주 번호'])
        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        copyShortcut = QShortcut(QKeySequence.Copy, self)
        pasteShortcut = QShortcut(QKeySequence.Paste, self)

        copyShortcut.activated.connect(self.copy)
        pasteShortcut.activated.connect(self.paste)

    #
    def scaleup(self):
        pass

    def scaledown(self):
        pass

    def copy(self):
        selectedRangeList = self.selectedRanges()
        if not selectedRangeList:
            return

        r = self.currentRow()
        c = self.currentColumn()
        selectedPoint = self.cellWidget(r, c)
        image = selectedPoint.imgpath

        QApplication.clipboard().setText(image)

    def paste(self):
        global NOT_SAVE
        image = QApplication.clipboard().text()
        r = self.currentRow()
        c = self.currentColumn()
        widget = TableWidget(image)
        self.setCellWidget(r, c, widget)
        NOT_SAVE = True

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls:
            e.accept()
        else:
            e.ignore()

    def dragMoveEvent(self, e):
        if e.mimeData().hasUrls:
            e.accept()
        else:
            e.ignore()

    def dropEvent(self, e):
        global NOT_SAVE
        e.setDropAction(QtCore.Qt.CopyAction)
        e.accept()
        if e.mimeData().hasUrls and FILE_NAME != "default":
            url = e.mimeData().urls()[0]
            url = str(url.toLocalFile())
            if url.split('.')[-1] == 'JPG' or url.split('.')[-1] == 'jpg':
                widget = TableWidget(url)
                row = self.currentRow()
                col = self.currentColumn()
                self.setCellWidget(row, col, widget)
                NOT_SAVE = True


class MyTabBar(QWidget):
    def __init__(self):
        super().__init__()
        self.table = MyTable()
        self.table.setEnabled(True)
        vbox = QVBoxLayout()
        vbox.addWidget(self.table)
        self.setLayout(vbox)


class WindowClass(QWidget, form_class):
    table = None
    wb = None
    sheet = None

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.buttonClick()

    def buttonClick(self):
        self.fileopen_btn.clicked.connect(self.openExcel)
        self.save_btn.clicked.connect(self.saveXpa)
        self.load_btn.clicked.connect(self.loadXpa)
        self.reload_btn.clicked.connect(self.reload)
        self.filesave_btn.clicked.connect(self.insert)
        self.delete_btn.clicked.connect(self.delete)
        self.deleteall_btn.clicked.connect(self.deleteall)
        self.scaleup_btn.clicked.connect(MyTable.scaleup)
        self.scaledown_btn.clicked.connect(MyTable.scaledown)

    def buttonActive(self):
        self.save_btn.setEnabled(True)
        self.reload_btn.setEnabled(True)
        self.filesave_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)
        self.deleteall_btn.setEnabled(True)
        self.load_btn.setEnabled(True)
        # self.scaleup_btn.setEnabled(True)
        # self.scaledown_btn.setEnabled(True)

    def deleteall(self):
        global NOT_SAVE
        msgBox = QMessageBox()
        msgBox.setWindowTitle('진짜 중요한 경고')
        msgBox.setIcon(QMessageBox.Critical)
        msgBox.setText('진짜 사진 다 지우겠습니까?')
        msgBox.setInformativeText('진짜 다 지우나요?')
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDefaultButton(QMessageBox.No)

        if msgBox.exec_() == QMessageBox.Yes:
            self.progressOn()
            tab_idx = myWindow.sheetlist.currentIndex()
            img = QPixmap()
            widget = TableWidget(None, pixmap=img)
            for r in range(self.sheetlist.widget(tab_idx).table.rowCount()):
                for c in range(self.sheetlist.widget(tab_idx).table.columnCount()):
                    self.sheetlist.widget(tab_idx).table.setCellWidget(r, c, widget)
            NOT_SAVE = True
            self.progressOff()

    def delete(self):
        global NOT_SAVE
        tab_idx = myWindow.sheetlist.currentIndex()
        r = self.sheetlist.widget(tab_idx).table.currentRow()
        c = self.sheetlist.widget(tab_idx).table.currentColumn()
        img = QPixmap()
        widget = TableWidget(None, pixmap=img)
        self.sheetlist.widget(tab_idx).table.setCellWidget(r, c, widget)
        NOT_SAVE = True

    def insert(self):
        global FILE_NAME
        print(FILE_NAME)
        if os.path.isfile(FILE_NAME):
            t = threading.Thread(target=self.inserting)
            t.start()
            print('파일있음')
        else:
            setLabelText(self.progress_lbl, f'[에러 ??] { FILE_NAME } 파일을 찾을 수 없습니다.')

    def inserting(self):
        global FILE_NAME, NOT_SAVE
        self.progressOn()
        column_list = ['A', 'I', 'Q']
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                if self.table.cellWidget(r, c):
                    row = 19 * r + 2
                    imgname = self.table.cellWidget(r, c).imgpath
                    insertinexcel(imgname, column_list[c], row, self.sheet)
        self.wb.save(FILE_NAME)
        setLabelText(self.progress_lbl, '[완료] 엑셀에 사진을 넣었습니다.')
        NOT_SAVE = False
        self.progressOff()

    def reload(self):
        tab_idx = myWindow.sheetlist.currentIndex()
        for r in range(self.sheetlist.widget(tab_idx).table.rowCount()):
            for c in range(self.sheetlist.widget(tab_idx).table.columnCount()):
                if self.sheetlist.widget(tab_idx).table.cellWidget(r, c):
                    imgpath = self.sheetlist.widget(tab_idx).table.cellWidget(r, c).imgpath
                    print(self.sheetlist.widget(tab_idx).table.cellWidget(r, c))
                    widget = TableWidget(imgpath)
                    self.sheetlist.widget(tab_idx).table.setCellWidget(r, c, widget)

    def loadExcel(self, fname):
        global FILE_NAME
        name = fname.split('/')[-1]
        FILE_NAME = fname
        # 엑셀 파일을 분석한다. openpyxl
        self.sheetlist.clear()
        try:
            self.wb = openpyxl.load_workbook(FILE_NAME)
        except FileNotFoundError:
            setLabelText(self.progress_lbl, '[에러 1] ' + name + ' 파일을 열지 못 했습니다. 엑셀 파일을 확인해주세요.')
            self.progressOff()
        else:
            name_list = self.wb.sheetnames
            for sheet_name in name_list:
                sheet = self.wb[sheet_name]
                row = (sheet.max_row - 1) // 19
                new_tab_bar = MyTabBar()
                new_tab_bar.table.setRowCount(row)
                self.sheetlist.addTab(new_tab_bar, sheet_name)
            setLabelText(self.fileopen_lbl, name)
            self.buttonActive()
            msgBox = QMessageBox()
            msgBox.setWindowTitle('엑셀 파일 불러오기 성공')
            msgBox.setIcon(QMessageBox.NoIcon)
            msgBox.setText(name + ' 파일을 불러왔습니다.')
            msgBox.setStandardButtons(QMessageBox.Yes)
            msgBox.setDefaultButton(QMessageBox.Yes)
            msgBox.exec_()
            setLabelText(self.progress_lbl, f'[완료] {FILE_NAME} 파일을 성공적으로 열었습니다.')
            self.progressOff()
            return True

    def openExcel(self):
        global NOT_SAVE
        if NOT_SAVE:
            msgBox = QMessageBox()
            msgBox.setWindowTitle('저장 안 됨')
            msgBox.setText('저장하지 않고 새 엑셀을 불러오겠습니까?')
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDefaultButton(QMessageBox.No)
            if msgBox.exec_() == QMessageBox.No:
                return

        setLabelText(self.progress_lbl, '[진행중] 파일을 여는 중...')
        self.progressOn()
        fname = QFileDialog.getOpenFileName(self, '엑셀 파일 선택', BASE_DIR, "Excel files (*.xlsx)")

        if fname[0]:
            if os.path.isfile(fname[0]):
                self.loadExcel(fname[0])
        else:
            setLabelText(self.progress_lbl, '[에러 0] 파일이 잘못 선택됐습니다. 다시 선택해주세요')
            self.progressOff()

    def saveXpa(self):
        """
        파일이름 QVarient
            ㄴ 탭 갯수
            ㄴ 탭 이름 QVarient + row 갯수 + col 갯수
                ㄴ 이미지정보 (xpa :  QVarient, xpae : QPixmap + QVariant)
        순으로 저장됨.
        :return:
        """
        global FILE_NAME, NOT_SAVE
        filters = "xpa File (*.xpa);; xpae File (*.xpae)"
        setLabelText(self.progress_lbl, '[진행중] 작업을 저장중...')
        self.progressOn()
        # 순수하게 *.xlsx 파일만 남기기 위해 작업
        excel_name = FILE_NAME.split('/')[-1]
        excel_name = excel_name.split("\\")[-1]
        # print(excel_name)
        save_file_name, _ = QFileDialog.getSaveFileName(self, "파일 저장하기", filter=filters)

        if save_file_name:
            save_file = QFile(save_file_name)
            save_file.open(QIODevice.WriteOnly)
            save_file.open(QIODevice.Append)
            stream_out = QDataStream(save_file)
            # 엑셀 파일 이름을 먼저 저장
            file_name = QVariant(excel_name)
            stream_out << file_name
            print('저장 파일명: ' + file_name.value())
            tab_num = QVariant(self.sheetlist.count())
            stream_out << tab_num

            if save_file_name.split('.')[1] == 'xpa':
                # 시트 이름을 저장
                for i in range(self.sheetlist.count()):
                    tab_str = QVariant(self.sheetlist.tabText(i))
                    t_str = tab_str.value() # test
                    stream_out << tab_str
                    row_num = QVariant(self.sheetlist.widget(i).table.rowCount())
                    t_str += ' ' + str(row_num.value()) # test
                    stream_out << row_num
                    col_num = QVariant(self.sheetlist.widget(i).table.columnCount())
                    t_str += ' ' + str(col_num.value()) # test
                    stream_out << col_num
                    print('저장 r, c: ' + t_str) # test
                    for r in range(self.sheetlist.widget(i).table.rowCount()):
                        for c in range(self.sheetlist.widget(i).table.columnCount()):   # 3
                            if self.sheetlist.widget(i).table.cellWidget(r, c):
                                output_str = QVariant(self.sheetlist.widget(i).table.cellWidget(r, c).imgpath)
                                stream_out << output_str
                            else:
                                output_str = QVariant('Null')
                                stream_out << output_str

            elif save_file_name.split('.')[1] == 'xpae':
                # 시트 이름을 저장
                for i in range(self.sheetlist.count()):
                    tab_str = QVariant(self.sheetlist.tabText(i))
                    stream_out << tab_str
                    row_num = QVariant(self.sheetlist.widget(i).table.rowCount())
                    stream_out << row_num
                    col_num = QVariant(self.sheetlist.widget(i).table.columnCount())
                    stream_out << col_num
                    for r in range(self.sheetlist.widget(i).table.rowCount()):
                        for c in range(self.sheetlist.widget(i).table.columnCount()):
                            if self.sheetlist.widget(i).table.cellWidget(r, c):
                                output_img = QPixmap(self.sheetlist.widget(i).table.cellWidget(r, c).img)
                                stream_out << output_img
                                output_str = QVariant(self.sheetlist.widget(i).table.cellWidget(r, c).imgpath)
                                stream_out << output_str
                            else:
                                output_img = QPixmap()
                                stream_out << output_img
                                output_str = QVariant('Null')
                                stream_out << output_str
            save_file.close()
            msgBox = QMessageBox()
            msgBox.setWindowTitle('파일 저장 성공')
            msgBox.setIcon(QMessageBox.NoIcon)
            msgBox.setText(save_file_name + ' 을 저장했습니다.')
            msgBox.setStandardButtons(QMessageBox.Yes)
            msgBox.setDefaultButton(QMessageBox.Yes)
            msgBox.exec_()
            NOT_SAVE = False
            setLabelText(self.progress_lbl, f'[완료] {FILE_NAME} 에 작업을 저장했습니다.')
        else:
            setLabelText(self.progress_lbl, '[에러 ??] 파일 저장에 실패했습니다.')
        self.progressOff()

    def loadXpa(self):
        """
            파일이름 QVarient
            ㄴ 탭 갯수
                ㄴ 탭 이름 QVarient + row 갯수 + col 갯수
                    ㄴ 이미지정보 (xpa :  QVarient, xpae : QPixmap)
            순으로 불러옴
            :return:
        """
        global FILE_NAME, SAVE_DIR, NOT_SAVE
        self.progressOn()
        if NOT_SAVE:
            msgBox = QMessageBox()
            msgBox.setWindowTitle('저장 안 됨')
            msgBox.setText('저장하지 않고 새 작업파일을 불러오겠습니까?')
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDefaultButton(QMessageBox.No)
            if msgBox.exec_() == QMessageBox.No:
                return
        filters = "xpa File (*.xpa);; xpae File (*.xpae)"
        load_file_name, _ = QFileDialog.getOpenFileName(self, '저장 파일 선택', BASE_DIR, filter=filters)
        # excel_name = FILE_NAME.split('/')[-1]

        # 순수하게 *.xlsx 파일만 남기기 위해 작업
        excel_name = FILE_NAME.split('/')[-1]
        excel_name = excel_name.split("\\")[-1]

        if load_file_name:
            load_file = QFile(load_file_name)
            load_file.open(QIODevice.ReadOnly)
            stream_in = QDataStream(load_file)
            file_name = QVariant()
            stream_in >> file_name
            print('로드 파일명: ' + file_name.value())
            # xpa 의 엑셀 파일이름과 설정한 엑셀 파일이름이 다를 경우
            if file_name.value() != excel_name:
                msgBox = QMessageBox()
                msgBox.setWindowTitle('에러')
                msgBox.setIcon(QMessageBox.NoIcon)
                msgBox.setText('저장 파일이 엑셀 파일의 저장이 아닙니다.')
                msgBox.setStandardButtons(QMessageBox.Yes)
                msgBox.setDefaultButton(QMessageBox.Yes)
                msgBox.exec_()
                setLabelText(self.progress_lbl, f'[에러] 저장 파일이 이 엑셀 파일의 저장이 아닙니다.')
                return
            # 파일 확장자가 xlsx가 아닐 경우 에러를 뿜뿜
            if file_name.value().split('.')[-1] != 'xlsx':
                print('파일이 xlsx가 아님')
                return
            FILE_NAME = BASE_DIR + '\\' + file_name.value()
            setLabelText(self.fileopen_lbl, file_name.value())
            print(FILE_NAME)
            tab_num = QVariant()
            stream_in >> tab_num
            tab_num = tab_num.value()

            for i in range(tab_num):
                tab_name = QVariant()
                stream_in >> tab_name
                row_num = QVariant()
                stream_in >> row_num
                col_num = QVariant()
                stream_in >> col_num
                new_tab_bar = MyTabBar()
                new_tab_bar.table.setRowCount(row_num.value())
                self.sheetlist.addTab(new_tab_bar, tab_name.value())
                for r in range(row_num.value()):
                    for c in range(col_num.value()):
                        if load_file_name.split('.')[1] == 'xpa':
                            input_str = QVariant()
                            stream_in >> input_str
                            if input_str.value() != 'Null':
                                widget = TableWidget(input_str.value())
                                self.sheetlist.widget(i).table.setCellWidget(r, c, widget)

                        elif load_file_name.split('.')[1] == 'xpae':
                            input_img = QPixmap()
                            stream_in >> input_img
                            input_str = QVariant()
                            stream_in >> input_str
                            widget = TableWidget(input_str.value(), pixmap=input_img)
                            self.sheetlist.widget(i).table.setCellWidget(r, c, widget)
            load_file.close()
            self.buttonActive()
            setLabelText(self.progress_lbl, f'[완료] {load_file_name} 파일을 불러왔습니다.')
            msgBox = QMessageBox()
            msgBox.setWindowTitle('파일 로드 성공')
            msgBox.setIcon(QMessageBox.NoIcon)
            msgBox.setText(load_file_name + ' 을 불러왔습니다.')
            msgBox.setStandardButtons(QMessageBox.Yes)
            msgBox.setDefaultButton(QMessageBox.Yes)
            msgBox.exec_()

        NOT_SAVE = False
        self.progressOff()

    def progressOn(self):
        self.progress_bar.setMaximum(0)

    def progressOff(self):
        self.progress_bar.setMaximum(1)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
